import os
import json
import shutil
import traceback

from kivy.app import App
from kivy.lang import Builder
from kivy.utils import platform
from kivy.clock import Clock
from kivy.properties import StringProperty, BooleanProperty, NumericProperty, DictProperty, ListProperty
from kivy.metrics import dp
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.recycleview import RecycleView
from kivy.uix.recycleboxlayout import RecycleBoxLayout
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.popup import Popup
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.scrollview import ScrollView
from kivy.uix.textinput import TextInput
from kivy.core.text import LabelBase

# --- 전역 설정 ---
SETTINGS_FILE = 'settings.json'
LOCAL_BASE = "/sdcard/Download/CheckSheet" if platform == 'android' else os.path.join(os.getcwd(), "CheckSheet_Data")

if os.path.exists("font.ttf"):
    try: LabelBase.register(name="Roboto", fn_regular="font.ttf")
    except: pass

# --- UI 디자인 ---
KV_UI = """
<Label>:
    font_name: 'Roboto'
<Button>:
    font_name: 'Roboto'
<TextInput>:
    font_name: 'Roboto'

<ListScreen>:
    BoxLayout:
        orientation: 'vertical'
        canvas.before:
            Color:
                rgba: 0.1, 0.1, 0.1, 1
            Rectangle:
                pos: self.pos
                size: self.size

        BoxLayout:
            size_hint_y: None
            height: '45dp'
            canvas.before:
                Color:
                    rgba: 0.15, 0.3, 0.4, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            Label:
                text: "Current File: " + app.current_filename
                bold: True

        BoxLayout:
            size_hint_y: None
            height: '60dp'
            padding: '5dp'
            spacing: '5dp'
            Button:
                text: 'Settings'
                on_release: app.open_smb_settings()
            Button:
                text: 'Excel'
                on_release: app.select_source('file')
            Button:
                text: 'PDF Folder'
                on_release: app.select_source('dir')
            Button:
                text: 'SAVE'
                background_color: 0.2, 0.7, 0.3, 1
                on_release: app.save_to_excel()

        BoxLayout:
            size_hint_y: None
            height: '40dp'
            canvas.before:
                Color:
                    rgba: 0.2, 0.2, 0.2, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            Button:
                text: 'No' + app.sort_indicator_no
                size_hint_x: 0.15
                on_release: app.sort_by('no')
            Button:
                text: 'Item Code' + app.sort_indicator_code
                size_hint_x: 0.4
                on_release: app.sort_by('item_code')
            Button:
                text: 'Qty' + app.sort_indicator_qty
                size_hint_x: 0.15
                on_release: app.sort_by('quantity')
            Button:
                text: 'Status'
                size_hint_x: 0.3

        CheckSheetRV:
            id: rv
            viewclass: 'RowWidget'
            RecycleBoxLayout:
                default_size: None, dp(65)
                default_size_hint: 1, None
                size_hint_y: None
                height: self.minimum_height
                orientation: 'vertical'
                spacing: '2dp'

<RowWidget>:
    orientation: 'horizontal'
    padding: [2, 2]
    canvas.before:
        Color:
            rgba: 0.2, 0.2, 0.2, 1
        Rectangle:
            pos: self.pos
            size: self.size
    Label:
        text: root.no
        size_hint_x: 0.15
    Button:
        text: root.item_code
        size_hint_x: 0.4
        background_normal: ''
        background_color: 0.2, 0.4, 0.6, 1
        on_release: root.open_pdf_external()
    Label:
        text: root.quantity
        size_hint_x: 0.15
    BoxLayout:
        size_hint_x: 0.3
        spacing: '2dp'
        Button:
            text: 'DONE'
            background_color: (0, 1, 0, 0.5) if root.complete else (0.3, 0.3, 0.3, 1)
            on_release: root.on_status('complete')
        Button:
            text: 'SHORT'
            background_color: (1, 1, 0, 0.5) if root.shortage else (0.3, 0.3, 0.3, 1)
            on_release: root.on_status('shortage')
        Button:
            text: 'REWK'
            background_color: (1, 0, 0, 0.5) if root.rework else (0.3, 0.3, 0.3, 1)
            on_release: root.on_status('rework')
"""

class ListScreen(Screen): pass
class CheckSheetRV(RecycleView): pass

class RowWidget(BoxLayout):
    no = StringProperty(''); item_code = StringProperty(''); quantity = StringProperty('')
    complete = BooleanProperty(False); shortage = BooleanProperty(False); rework = BooleanProperty(False)
    
    def on_status(self, st):
        App.get_running_app().update_item_status(self.item_code, self.no, st)
    
    def open_pdf_external(self):
        App.get_running_app().open_pdf_in_external_app(self.item_code)

class CheckSheetApp(App):
    excel_path = StringProperty(''); pdf_folder_path = StringProperty('')
    current_filename = StringProperty('Please select file')
    smb_config = DictProperty({'ip': '', 'user': '', 'pass': ''})
    sort_indicator_no = StringProperty(''); sort_indicator_code = StringProperty(''); sort_indicator_qty = StringProperty('')
    sort_states = {}

    def build(self):
        if platform == 'android':
            try:
                from kivy.core.window import Window
                Window.softinput_mode = 'below_target'
            except: pass
        self.load_settings()
        Builder.load_string(KV_UI)
        sm = ScreenManager()
        sm.add_widget(ListScreen(name='list'))
        return sm

    def on_start(self):
        Clock.schedule_once(self.delayed_init, 1)

    def delayed_init(self, dt):
        self.ask_permissions()
        if self.excel_path and os.path.exists(self.excel_path):
            self.load_excel_data(self.excel_path)

    def ask_permissions(self):
        if platform != 'android': return
        try:
            from android.permissions import request_permissions, Permission
            from jnius import autoclass
            request_permissions([Permission.READ_EXTERNAL_STORAGE, Permission.WRITE_EXTERNAL_STORAGE, Permission.INTERNET])
            Env = autoclass('android.os.Environment')
            if hasattr(Env, 'isExternalStorageManager') and not Env.isExternalStorageManager():
                mActivity = autoclass('org.kivy.android.PythonActivity').mActivity
                Intent = autoclass('android.content.Intent')
                Settings = autoclass('android.provider.Settings')
                Uri = autoclass('android.net.Uri').fromParts("package", mActivity.getPackageName(), None)
                intent = Intent(Settings.ACTION_MANAGE_APP_ALL_FILES_ACCESS_PERMISSION)
                intent.setData(uri); mActivity.startActivity(intent)
        except: pass

    # --- 핵심: 외부 PDF 앱 연동 ---
    def open_pdf_in_external_app(self, item_code):
        base_path = self.pdf_folder_path if self.pdf_folder_path else LOCAL_BASE
        pdf_path = os.path.join(base_path, f"{item_code}.pdf")
        
        if not os.path.exists(pdf_path):
            self.show_popup("Notice", f"File not found:\n{item_code}.pdf")
            return

        if platform == 'android':
            try:
                from jnius import autoclass, cast
                mActivity = autoclass('org.kivy.android.PythonActivity').mActivity
                Intent = autoclass('android.content.Intent')
                Uri = autoclass('android.net.Uri')
                File = autoclass('java.io.File')
                
                # Use StrictMode bypass to open file:// URI
                StrictMode = autoclass('android.os.StrictMode')
                StrictMode.disableDeathOnFileUriExposure()
                
                file = File(pdf_path)
                uri = Uri.fromFile(file)
                
                intent = Intent(Intent.ACTION_VIEW)
                intent.setDataAndType(uri, "application/pdf")
                intent.addFlags(Intent.FLAG_ACTIVITY_NEW_TASK)
                intent.addFlags(Intent.FLAG_GRANT_READ_URI_PERMISSION)
                
                chooser = Intent.createChooser(intent, cast('java.lang.CharSequence', "Open with..."))
                mActivity.startActivity(chooser)
            except Exception as e:
                self.show_popup("Error", f"Failed to open external app: {e}")
        else:
            self.show_popup("Notice", f"External app not supported on Windows.\nPath: {pdf_path}")

    def load_excel_data(self, path):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True); ws = wb.active; rows = list(ws.rows)
            if not rows: return
            h = [str(c.value).strip().lower() if c.value else "" for c in rows[0]]
            idx_no, idx_code, idx_qty = h.index('no'), h.index('품목코드'), h.index('수량')
            rv_data = []
            for i, row in enumerate(rows[1:]):
                if not row[idx_code].value: continue
                rv_data.append({
                    'no': str(row[idx_no].value or ''), 'item_code': str(row[idx_code].value or ''), 'quantity': str(row[idx_qty].value or ''),
                    'complete': str(ws.cell(row=i+2, column=h.index('완료')+1).value or '').upper() == 'V' if '완료' in h else False,
                    'shortage': str(ws.cell(row=i+2, column=h.index('수량부족')+1).value or '').upper() == 'V' if '수량부족' in h else False,
                    'rework': str(ws.cell(row=i+2, column=h.index('재작업')+1).value or '').upper() == 'V' if '재작업' in h else False,
                    'real_index': i
                })
            self.root.get_screen('list').ids.rv.data = rv_data
            self.current_filename = os.path.basename(path)
        except Exception as e: self.show_popup("Load Failed", str(e))

    def save_to_excel(self):
        if not self.excel_path: return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.excel_path); ws = wb.active; h = [str(c.value).strip() if c.value else "" for c in ws[1]]
            target = ['완료', '수량부족', '재작업']; cols = {}
            for n in target:
                if n in h: cols[n] = h.index(n) + 1
                else:
                    new_idx = len(h) + 1; ws.cell(row=1, column=new_idx).value = n; cols[n] = new_idx; h.append(n)
            for d in self.root.get_screen('list').ids.rv.data:
                r = d['real_index'] + 2
                ws.cell(row=r, column=cols['완료']).value = 'V' if d.get('complete') else ''
                ws.cell(row=r, column=cols['수량부족']).value = 'V' if d.get('shortage') else ''
                ws.cell(row=r, column=cols['재작업']).value = 'V' if d.get('rework') else ''
            wb.save(self.excel_path); self.show_popup("Notice", "Save Successful")
        except Exception as e: self.show_popup("Save Failed", str(e))

    def update_item_status(self, ic, no, st):
        rv = self.root.get_screen('list').ids.rv
        for d in rv.data:
            if d['item_code'] == ic and d['no'] == no:
                if st == 'complete':
                    d['complete'] = not d['complete']
                    if d['complete']: d['shortage'] = d['rework'] = False
                elif st == 'shortage':
                    d['shortage'] = not d['shortage']
                    if d['shortage']: d['complete'] = d['rework'] = False
                elif st == 'rework':
                    d['rework'] = not d['rework']
                    if d['rework']: d['complete'] = d['shortage'] = False
                break
        rv.refresh_from_data()

    def sort_by(self, col):
        rv = self.root.get_screen('list').ids.rv
        if not rv.data: return
        new_s = 'desc' if self.sort_states.get(col) == 'asc' else 'asc'
        self.sort_states = {col: new_s}
        for k in ['no','code','qty']: setattr(self, f'sort_indicator_{k}', '')
        setattr(self, f'sort_indicator_{col if col != "item_code" else "code"}', " ▲" if new_s == 'asc' else " ▼")
        def sk(x):
            v = str(x.get(col, '')).lower()
            if col in ['no', 'quantity']:
                try: return int(''.join(filter(str.isdigit, v)) or 0)
                except: return v
            return v
        rv.data = sorted(rv.data, key=sk, reverse=(new_s == 'desc')); rv.refresh_from_data()

    def select_source(self, mode):
        content = BoxLayout(orientation='vertical', padding=20, spacing=20)
        pop = Popup(title="Select", content=content, size_hint=(0.8, 0.4))
        content.add_widget(Button(text="Phone Local", on_release=lambda x: (pop.dismiss(), self.open_local_browser(mode))))
        content.add_widget(Button(text="PC Shared Folder", on_release=lambda x: (pop.dismiss(), self.open_smb_shares_browser(mode))))
        pop.open()

    def open_local_browser(self, mode):
        fc = FileChooserListView(path="/storage/emulated/0" if platform=='android' else os.getcwd())
        if mode == 'dir': fc.dirselect = True
        content = BoxLayout(orientation='vertical', padding=5); content.add_widget(fc)
        pop = Popup(title="Choose File", content=content, size_hint=(0.9, 0.9))
        def confirm(x):
            t = fc.selection[0] if fc.selection else fc.path
            if mode == 'file' and os.path.isfile(t): self.excel_path = t; self.load_excel_data(t); self.save_settings(); pop.dismiss()
            elif mode == 'dir' and os.path.isdir(t): self.pdf_folder_path = t; self.save_settings(); pop.dismiss()
        content.add_widget(Button(text="Select", size_hint_y=None, height=60, on_release=confirm)); pop.open()

    def open_smb_shares_browser(self, mode):
        conn = self.get_smb_conn_only()
        if not conn: self.show_popup("Notice", "SMB Failed"); return
        content = BoxLayout(orientation='vertical'); lb = BoxLayout(orientation='vertical', size_hint_y=None); lb.bind(minimum_height=lb.setter('height')); scroll = ScrollView(); scroll.add_widget(lb)
        pop = Popup(title="SMB Shares", content=content, size_hint=(0.9, 0.9))
        for s in conn.listShares():
            if s.isSpecial or s.name.endswith('$'): continue
            b = Button(text=s.name, size_hint_y=None, height=80); b.bind(on_release=lambda x, n=s.name: self.open_smb_files_browser(conn, n, "/", mode, pop)); lb.add_widget(b)
        content.add_widget(scroll); pop.open()

    def open_smb_files_browser(self, conn, share, path, mode, parent):
        content = BoxLayout(orientation='vertical'); lb = BoxLayout(orientation='vertical', size_hint_y=None); lb.bind(minimum_height=lb.setter('height')); scroll = ScrollView(); scroll.add_widget(lb)
        pop = Popup(title=f"SMB: {share}", content=content, size_hint=(0.9, 0.9))
        def refresh(cp):
            lb.clear_widgets()
            for f in conn.listPath(share, cp):
                if f.filename in ['.', '..']: continue
                b = Button(text=f.filename, size_hint_y=None, height=80)
                def click(x, file=f):
                    np = os.path.join(cp, file.filename).replace("\\", "/")
                    if file.isDirectory: self.open_smb_files_browser(conn, share, np, mode, parent)
                    elif mode == 'file':
                        local = os.path.join(LOCAL_BASE, file.filename)
                        if not os.path.exists(LOCAL_BASE): os.makedirs(LOCAL_BASE)
                        with open(local, 'wb') as lf: conn.retrieveFile(share, np, lf)
                        self.excel_path = local; self.load_excel_data(local); self.save_settings(); pop.dismiss(); parent.dismiss()
                b.bind(on_release=click); lb.add_widget(b)
        refresh(path); content.add_widget(scroll); pop.open()

    def get_smb_conn_only(self):
        try:
            from smb.SMBConnection import SMBConnection
            c = self.smb_config; conn = SMBConnection(c['user'], c['pass'], "App", c['ip'], use_ntlm_v2=True, is_direct_tcp=True)
            if conn.connect(c['ip'], 445, timeout=3): return conn
        except: return None

    def open_smb_settings(self):
        content = BoxLayout(orientation='vertical', padding=10)
        ips = TextInput(text=self.smb_config['ip'], multiline=False); usr = TextInput(text=self.smb_config['user'], multiline=False); pas = TextInput(text=self.smb_config['pass'], password=True, multiline=False)
        content.add_widget(Label(text="IP")); content.add_widget(ips); content.add_widget(Label(text="ID")); content.add_widget(usr); content.add_widget(Label(text="PW")); content.add_widget(pas)
        pop = Popup(title="SMB Config", content=content, size_hint=(0.8, 0.6))
        def save(x): self.smb_config = {'ip':ips.text,'user':usr.text,'pass':pas.text}; self.save_settings(); pop.dismiss()
        content.add_widget(Button(text="SAVE", on_release=save)); pop.open()

    def load_settings(self):
        if os.path.exists(SETTINGS_FILE):
            try:
                with open(SETTINGS_FILE, 'r') as f:
                    d = json.load(f); self.excel_path = d.get('excel_path', ''); self.pdf_folder_path = d.get('pdf_folder_path', ''); self.smb_config = d.get('smb_config', {'ip':'','user':'','pass':''})
            except: pass

    def save_settings(self):
        with open(SETTINGS_FILE, 'w') as f: json.dump({'excel_path': self.excel_path, 'pdf_folder_path': self.pdf_folder_path, 'smb_config': self.smb_config}, f)

    def show_popup(self, title, msg): Popup(title=title, content=Label(text=str(msg)), size_hint=(0.8, 0.4)).open()

if __name__ == '__main__':
    CheckSheetApp().run()
